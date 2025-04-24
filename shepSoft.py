import os
import platform
import csv

def banner():
    print("""
    ███████╗██╗  ██╗███████╗██████╗ 
    ██╔════╝██║  ██║██╔════╝██╔══██╗
    ██████╗ ███████║█████╗  ██████╔╝
    ╚═══██║ ██╔══██║██╔══╝  ██╔═══╝ 
   ███████║ ██║  ██║███████╗██║     
    ╚═════╝ ╚═╝  ╚═╝╚══════╝╚═╝
     
      tg - @PA3O4EPOBAJI_BCEX
    """)

def clear_console():
    os.system('cls' if platform.system() == 'Windows' else 'clear')

def get_directory_size(directory):
    total_size = 0
    for dirpath, dirnames, filenames in os.walk(directory):
        for f in filenames:
            fp = os.path.join(dirpath, f)
            total_size += os.path.getsize(fp)
    return total_size

def format_size(size):
    if size < 1024 ** 2:  # меньше 1 МБ
        return f"{size} байт"
    elif size < 1024 ** 3:  # меньше 1 ГБ
        return f"{size / 1024 ** 2:.2f} МБ"
    else:  # 1 ГБ и больше
        return f"{size / 1024 ** 3:.2f} ГБ"

def search_in_file(file_path, keyword):
    try:
        if file_path.endswith('.csv'):
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                for row in reader:
                    if any(keyword in str(cell) for cell in row):
                        print(f"Found in {file_path}: {row}")
        
        elif file_path.endswith('.xlsx') or file_path.endswith('.xls'):
            # Обработка Excel файлов через openpyxl
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    if any(keyword in str(cell) for cell in row):
                        print(f"Found in {file_path} ({sheet}): {row}")

        elif file_path.endswith('.txt') or file_path.endswith('.sql'):
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line in f:
                    if keyword in line:
                        print(f"Found in {file_path}: {line.strip()}")
    
    except Exception as e:
        print(f"Error processing {file_path}: {e}")

def search_keyword_in_directory(directory, keyword):
    for filename in os.listdir(directory):
        file_path = os.path.join(directory, filename)
        if os.path.isfile(file_path):
            search_in_file(file_path, keyword)

def main():
    while True:
        clear_console()
        banner()
        print("\nМеню:")
        
        # Показать размер папки с базами данных
        db_directory = 'bases'
        db_size = get_directory_size(db_directory)
        formatted_size = format_size(db_size)
        print(f"Размер папки с базами данных: {formatted_size}")

        print("1. Поиск в базах данных")
        print("2. Выход")
        choice = input("Выберите пункт меню (1-2): ").strip()
        
        if choice == '1':
            print("Введите ключевое слово для поиска:")
            keyword = input().strip()
            print("Поиск в папке 'bases'...")
            search_keyword_in_directory(db_directory, keyword)
            input("Нажмите Enter, чтобы продолжить...")
        elif choice == '2':
            print("Выход из программы.")
            break
        else:
            print("Неверный выбор. Пожалуйста, выберите 1 или 2.")
            input("Нажмите Enter, чтобы продолжить...")

if __name__ == "__main__":
    main()